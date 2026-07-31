"""导出服务（FR-06）：抽取结果 JSON / Excel / Markdown 三种格式。"""

from __future__ import annotations

import html
import io
import json
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.storage import db


def _display(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _md_escape(text: str) -> str:
    return str(text).replace("|", "\\|").replace("\n", " ")


def build_export(row, fmt: str) -> tuple[bytes, str, str]:
    """构建导出文件，返回 (内容字节, media_type, 文件名)。"""
    schema = db.get_schema(row["schema_id"])
    fields = {
        f["key"]: f for f in (db.jloads(schema["fields_json"], []) if schema else [])
    }
    data = db.jloads(row["data_json"], {})
    confidence = db.jloads(row["confidence_json"], {})
    field_status = db.jloads(row["field_status_json"], {})
    citations = db.jloads(row["citations_json"], {})

    rows: list[dict[str, Any]] = []
    for key, value in data.items():
        field = fields.get(key, {})
        cites = citations.get(key) or []
        snippets = [
            str(c.get("snippet") or "")
            for c in cites
            if isinstance(c, dict) and c.get("snippet")
        ]
        rows.append(
            {
                "key": key,
                "label": field.get("label") or key,
                "value": _display(value),
                "confidence": confidence.get(key),
                "status": field_status.get(key, ""),
                "citation": "；".join(snippets)[:300],
            }
        )

    ext_id = row["id"]
    schema_name = schema["name"] if schema else ""
    if fmt == "json":
        payload = {
            "id": ext_id,
            "doc_id": row["doc_id"],
            "schema_id": row["schema_id"],
            "schema_name": schema_name,
            "status": row["status"],
            "source": row["source"],
            "confirmed_at": row["confirmed_at"],
            "created_at": row["created_at"],
            "rows": rows,
            "data": data,
            "confidence": confidence,
            "field_status": field_status,
        }
        content = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
        return content, "application/json; charset=utf-8", f"extraction-{ext_id}.json"

    if fmt == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "抽取结果"
        header = ["字段", "标签", "值", "置信度", "状态", "依据"]
        ws.append(header)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for r in rows:
            ws.append(
                [
                    r["key"], r["label"], r["value"],
                    r["confidence"], r["status"], r["citation"],
                ]
            )
        for i, width in enumerate((20, 20, 40, 12, 12, 60), start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
        buf = io.BytesIO()
        wb.save(buf)
        media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return buf.getvalue(), media, f"extraction-{ext_id}.xlsx"

    # markdown
    lines = [
        f"# 抽取结果 #{ext_id}",
        "",
        f"- 文档 ID：{row['doc_id']}",
        f"- Schema：{schema_name}",
        f"- 状态：{row['status']}（来源：{row['source']}）",
        f"- 确认时间：{row['confirmed_at'] or '-'}",
        "",
        "| 字段 | 标签 | 值 | 置信度 | 状态 | 依据 |",
        "| --- | --- | --- | --- | --- | --- |",
    ]
    for r in rows:
        lines.append(
            f"| {_md_escape(r['key'])} | {_md_escape(r['label'])} | "
            f"{_md_escape(r['value'])} | {r['confidence']} | {r['status']} | "
            f"{_md_escape(r['citation'])} |"
        )
    content = ("\n".join(lines) + "\n").encode("utf-8")
    return content, "text/markdown; charset=utf-8", f"extraction-{ext_id}.md"

def _fmt_delta(delta: float | None) -> str:
    if delta is None:
        return "-"
    return f"{delta:+.1f}%"


def build_compare_export(row, fmt: str) -> tuple[bytes, str, str]:
    """构建对比报告（md/html），返回 (内容字节, media_type, 文件名)。"""
    result = db.jloads(row["result_json"], {})
    doc_a = db.get_document(row["doc_a_id"])
    doc_b = db.get_document(row["doc_b_id"])
    schema = db.get_schema(row["schema_id"])
    field_diff = result.get("field_diff", [])
    section_diff = result.get("section_diff", [])
    summary = result.get("summary", "")
    name_a = doc_a["original_name"] if doc_a else f"文档{row['doc_a_id']}"
    name_b = doc_b["original_name"] if doc_b else f"文档{row['doc_b_id']}"
    schema_name = schema["name"] if schema else ""
    compare_id = row["id"]

    if fmt == "html":
        lines = [
            "<!DOCTYPE html>",
            "<html lang='zh-CN'><head><meta charset='utf-8'>",
            f"<title>文档对比报告 #{compare_id}</title>",
            (
                "<style>body{font-family:'Microsoft YaHei',sans-serif;max-width:960px;"
                "margin:24px auto;color:#222}"
                "table{border-collapse:collapse;width:100%;margin:12px 0}"
                "th,td{border:1px solid #ccc;padding:8px;text-align:left;font-size:14px}"
                "th{background:#f3f4f6}"
                ".badge{display:inline-block;padding:2px 8px;border-radius:10px;font-size:12px}"
                ".same{background:#e8f5e9}.changed{background:#fff3e0}"
                ".only_a{background:#e3f2fd}.only_b{background:#fce4ec}.both_missing{background:#eee}"
                ".added{background:#e8f5e9}.removed{background:#fce4ec}</style></head><body>"
            ),
            f"<h1>文档对比报告 #{compare_id}</h1>",
            (
                f"<p><b>{html.escape(name_a)}</b> vs <b>{html.escape(name_b)}</b>"
                f"（Schema：{html.escape(schema_name)}，生成于 {row['created_at']}）</p>"
            ),
            f"<h2>摘要</h2><p>{html.escape(summary)}</p>",
            "<h2>字段差异</h2>",
            "<table><tr><th>字段</th><th>标签</th><th>文档 A</th><th>文档 B</th><th>状态</th><th>变化</th></tr>",
        ]
        for d in field_diff:
            status = d.get("status", "")
            lines.append(
                f"<tr><td>{html.escape(str(d.get('key', '')))}</td>"
                f"<td>{html.escape(str(d.get('label', '')))}</td>"
                f"<td>{html.escape(_display(d.get('value_a')))}</td>"
                f"<td>{html.escape(_display(d.get('value_b')))}</td>"
                f"<td><span class='badge {status}'>{status}</span></td>"
                f"<td>{_fmt_delta(d.get('delta_pct'))}</td></tr>"
            )
        lines.append("</table><h2>章节差异</h2>")
        lines.append("<table><tr><th>章节</th><th>状态</th><th>相似度</th></tr>")
        for d in section_diff:
            status = d.get("status", "")
            lines.append(
                f"<tr><td>{html.escape(str(d.get('title', '')))}</td>"
                f"<td><span class='badge {status}'>{status}</span></td>"
                f"<td>{d.get('similarity', '')}</td></tr>"
            )
        lines.append("</table></body></html>")
        content = ("\n".join(lines) + "\n").encode("utf-8")
        return content, "text/html; charset=utf-8", f"compare-{compare_id}.html"

    lines = [
        f"# 文档对比报告 #{compare_id}",
        "",
        f"- 文档 A：{_md_escape(name_a)}",
        f"- 文档 B：{_md_escape(name_b)}",
        f"- Schema：{_md_escape(schema_name)}",
        f"- 生成时间：{row['created_at']}",
        "",
        "## 摘要",
        "",
        summary or "-",
        "",
        "## 字段差异",
        "",
        "| 字段 | 标签 | 文档 A | 文档 B | 状态 | 变化 |",
        "| --- | --- | --- | --- | --- | --- |",
    ]
    for d in field_diff:
        lines.append(
            f"| {_md_escape(d.get('key', ''))} | {_md_escape(d.get('label', ''))} | "
            f"{_md_escape(_display(d.get('value_a')))} | {_md_escape(_display(d.get('value_b')))} | "
            f"{d.get('status', '')} | {_fmt_delta(d.get('delta_pct'))} |"
        )
    lines += ["", "## 章节差异", "", "| 章节 | 状态 | 相似度 |", "| --- | --- | --- |"]
    for d in section_diff:
        lines.append(
            f"| {_md_escape(d.get('title', ''))} | {d.get('status', '')} | {d.get('similarity', '')} |"
        )
    content = ("\n".join(lines) + "\n").encode("utf-8")
    return content, "text/markdown; charset=utf-8", f"compare-{compare_id}.md"
