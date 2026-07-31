"""文档解析服务：按格式分派解析器，输出结构化页流。

页流结构：
    pages: list[{page: int, blocks: [{kind, text|table|image_path, heading_level}]}]
"""
from __future__ import annotations

import io
from pathlib import Path
from typing import Any

from app.storage import files as file_store
from app.utils.logging import get_logger
from app.utils.text import clean_text

logger = get_logger("parser")


class ParseError(RuntimeError):
    """解析失败（用户可读信息）。"""


class OcrUnavailable(ParseError):
    def __init__(self, message: str = ""):
        super().__init__(message or "扫描件需要 OCR：请安装 PaddleOCR（pip install -r requirements-ocr.txt）或配置视觉模型")


def parse_document(path: Path, ext: str) -> dict[str, Any]:
    """入口：按扩展名分派解析器。返回 {pages, page_count, char_count, warnings}。"""
    ext = ext.lower()
    if ext == ".pdf":
        pages, warnings = _parse_pdf(path)
    elif ext == ".docx":
        pages, warnings = _parse_docx(path)
    elif ext == ".xlsx":
        pages, warnings = _parse_xlsx(path)
    elif ext in {".png", ".jpg", ".jpeg", ".webp"}:
        pages, warnings = _parse_image(path, ext)
    else:
        raise ParseError(f"不支持的格式 {ext}")
    char_count = sum(len(b.get("text", "") or b.get("table", "") or "") for p in pages for b in p["blocks"])
    return {"pages": pages, "page_count": len(pages), "char_count": char_count, "warnings": warnings}


# ------------------------------------------------------------------ PDF

def _parse_pdf(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    import pdfplumber

    pages: list[dict[str, Any]] = []
    warnings: list[str] = []
    with pdfplumber.open(path) as pdf:
        for idx, page in enumerate(pdf.pages):
            blocks: list[dict[str, Any]] = []
            # 1) 图片提取（先于文本，图片块独立）
            for img in page.images:
                try:
                    cropped = page.crop((img["x0"], img["top"], img["x1"], img["bottom"]))
                    buf = io.BytesIO()
                    cropped.to_image(resolution=120).save(buf, format="PNG")
                    name = file_store.save_image_bytes(
                        f"doc_{path.stem[:20]}_{idx+1}_{img['x0']:.0f}_{img['top']:.0f}.png", buf.getvalue())
                    blocks.append({"kind": "image", "image_path": name, "heading_level": None})
                except Exception as e:  # noqa: BLE001
                    warnings.append(f"第 {idx+1} 页图片提取失败：{e}")
            # 2) 表格提取（转为 Markdown）
            try:
                tables = page.extract_tables()
            except Exception:  # noqa: BLE001
                tables = []
            for tb in tables:
                md = _table_to_markdown(tb)
                if md:
                    blocks.append({"kind": "table", "table": md, "heading_level": None})
            # 3) 文本行 → 标题/正文块（按连续行聚合）
            try:
                lines = page.extract_text_lines() or []
            except Exception:  # noqa: BLE001
                lines = []
            if lines:
                sizes = [ln.get("size", 0) or 0 for ln in lines if (ln.get("text") or "").strip()]
                max_size = max(sizes) if sizes else 0
                buf_text: list[str] = []
                for ln in lines:
                    text = clean_text(ln.get("text", ""))
                    if not text:
                        continue
                    size = ln.get("size", 0) or 0
                    if _is_heading(size, max_size, text):
                        if buf_text:
                            blocks.append({"kind": "text", "text": "\n".join(buf_text), "heading_level": None})
                            buf_text = []
                        blocks.append({"kind": "text", "text": text, "heading_level": _heading_level(size, max_size)})
                    else:
                        buf_text.append(text)
                if buf_text:
                    blocks.append({"kind": "text", "text": "\n".join(buf_text), "heading_level": None})
            # 乱码检测
            page_text = "".join(ln.get("text", "") for ln in lines)
            if page_text and page_text.count("\ufffd") / max(len(page_text), 1) > 0.05:
                warnings.append(f"第 {idx+1} 页疑似乱码，已跳过")
                continue
            pages.append({"page": idx + 1, "blocks": blocks})
    if not pages:
        raise ParseError("PDF 未解析出任何内容（可能是扫描件且未安装 OCR）")
    return pages, warnings


def _is_heading(size: float, max_size: float, text: str) -> bool:
    return size >= max_size - 0.6 and len(text) <= 80


def _heading_level(size: float, max_size: float) -> int:
    diff = max_size - size
    if diff < 0.3:
        return 1
    if diff < 1.2:
        return 2
    return 3


def _table_to_markdown(table: list[list[Any]]) -> str:
    if not table:
        return ""
    rows = [[clean_text(str(c)) if c is not None else "" for c in row] for row in table]
    rows = [r for r in rows if any(r)]
    if not rows:
        return ""
    width = max(len(r) for r in rows)
    rows = [r + [""] * (width - len(r)) for r in rows]
    header = rows[0]
    body = rows[1:] or rows  # 无表头时首行作为表头
    lines = ["| " + " | ".join(header) + " |", "| " + " | ".join(["---"] * width) + " |"]
    lines += ["| " + " | ".join(r) + " |" for r in body]
    return "\n".join(lines)


# ------------------------------------------------------------------ DOCX

def _parse_docx(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    from docx import Document

    doc = Document(path)
    blocks: list[dict[str, Any]] = []
    for item in _iter_docx_items(doc):
        if item["kind"] == "paragraph":
            text = clean_text(item["obj"].text)
            if not text:
                continue
            level = _docx_heading_level(item["obj"])
            if level:
                blocks.append({"kind": "text", "text": text, "heading_level": level})
            else:
                blocks.append({"kind": "text", "text": text, "heading_level": None})
        else:
            md = _table_to_markdown([[c.text for c in row.cells] for row in item["obj"].rows])
            if md:
                blocks.append({"kind": "table", "table": md, "heading_level": None})
    return [{"page": 1, "blocks": blocks}], []


def _iter_docx_items(doc):
    from docx.oxml.ns import qn
    from docx.table import Table
    from docx.text.paragraph import Paragraph

    body = doc.element.body
    for child in body.iterchildren():
        if child.tag == qn("w:p"):
            yield {"kind": "paragraph", "obj": Paragraph(child, doc)}
        elif child.tag == qn("w:tbl"):
            yield {"kind": "table", "obj": Table(child, doc)}


def _docx_heading_level(paragraph) -> int | None:
    style = (paragraph.style.name or "").lower()
    for i in (1, 2, 3):
        if f"heading {i}" in style or f"标题 {i}" in style:
            return i
    return None


# ------------------------------------------------------------------ XLSX

def _parse_xlsx(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    blocks: list[dict[str, Any]] = []
    warnings: list[str] = []
    for ws in wb.worksheets:
        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            vals = [clean_text(str(c)) if c is not None else "" for c in row]
            if any(vals):
                rows.append(vals)
            if len(rows) >= 500:
                break
        if rows:
            md = _table_to_markdown(rows)
            blocks.append({"kind": "table", "table": md, "heading_level": None})
        elif ws.max_row and ws.max_row > 0:
            warnings.append(f"工作表「{ws.title}」无数据")
    wb.close()
    if not blocks:
        raise ParseError("Excel 未解析出任何内容")
    return [{"page": 1, "blocks": blocks}], warnings


# ------------------------------------------------------------------ 图片 / OCR

def _parse_image(path: Path, ext: str) -> tuple[list[dict[str, Any]], list[str]]:
    try:
        from paddleocr import PaddleOCR  # type: ignore
    except ImportError as e:
        raise OcrUnavailable() from e
    try:
        ocr = PaddleOCR(use_angle_cls=True, lang="ch", show_log=False)
        result = ocr.ocr(str(path), cls=True)
    except Exception as e:
        raise ParseError(f"OCR 识别失败：{e}") from e
    lines: list[str] = []
    for page_result in result or []:
        for item in page_result or []:
            text = clean_text(item[1][0] if item and len(item) > 1 else "")
            if text:
                lines.append(text)
    if not lines:
        raise ParseError("OCR 未识别出文本")
    blocks = [{"kind": "text", "text": "\n".join(lines), "heading_level": None}]
    return [{"page": 1, "blocks": blocks}], []
