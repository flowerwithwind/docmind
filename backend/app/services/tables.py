"""表格转换服务（B2）：Excel/PDF 表格块 → 可查询的行列结构。

- xlsx_to_table：openpyxl 读取 Excel 首个工作表（首行为表头）；
- markdown_to_table：把 PDF/DOCX 解析产出的 Markdown 表格块清洗归一为行列结构；
- register_from_doc：把已解析文档中的 table 块批量注册为可查询表。
"""
from __future__ import annotations

import io
import re
from pathlib import Path
from typing import Any

from app.services.table_store import MAX_TABLE_ROWS, TableError, table_store
from app.storage import db
from app.utils.logging import get_logger
from app.utils.text import clean_text

logger = get_logger("tables")

_SEP_CELL_RE = re.compile(r"^:?-{2,}:?$")


def xlsx_to_table(content: bytes | str | Path) -> dict[str, Any]:
    """读取 Excel 首个工作表（首行为表头），返回 {columns, rows, sheet}。"""
    from openpyxl import load_workbook

    if isinstance(content, (str, Path)):
        wb = load_workbook(content, read_only=True, data_only=True)
    else:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0] if wb.worksheets else None
        if ws is None:
            raise TableError("Excel 中没有工作表")
        raw: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            vals = [clean_text(str(c)) if c is not None else "" for c in row]
            if any(vals):
                raw.append(vals)
            if len(raw) >= MAX_TABLE_ROWS + 1:
                break
        if not raw:
            raise TableError(f"工作表「{ws.title}」无数据")
        columns = [c or f"列{i + 1}" for i, c in enumerate(raw[0])]
        return {"columns": columns, "rows": raw[1:], "sheet": ws.title}
    finally:
        wb.close()


def _is_sep_row(row: list[str]) -> bool:
    nonempty = [c for c in row if c]
    return bool(nonempty) and all(_SEP_CELL_RE.match(c) for c in nonempty)


def markdown_to_table(md: str) -> dict[str, Any]:
    """把 Markdown 表格文本清洗归一为 {columns, rows}。

    兼容 PDF/DOCX/XLSX 解析块（parser._table_to_markdown 产物）：
    跳过分隔行、剔除空行、补齐行宽、压缩空白。
    """
    lines = [ln.strip() for ln in (md or "").splitlines() if ln.strip().startswith("|")]
    if not lines:
        raise TableError("未识别到 Markdown 表格行")
    grid: list[list[str]] = []
    for ln in lines:
        cells = [clean_text(c) for c in ln.strip().strip("|").split("|")]
        if any(cells):
            grid.append(cells)
    if not grid:
        raise TableError("Markdown 表格为空")
    width = max(len(r) for r in grid)
    grid = [r + [""] * (width - len(r)) for r in grid]
    if _is_sep_row(grid[0]) and len(grid) > 1:
        grid = grid[1:]  # 防御：首行即分隔行时丢弃
    header = [c or f"列{i + 1}" for i, c in enumerate(grid[0])]
    rows = [r for r in grid[1:] if not _is_sep_row(r) and any(r)]
    return {"columns": header, "rows": rows}


def register_xlsx(content: bytes | str | Path, name: str = "Excel 表格", source: str = "xlsx") -> Any:
    """Excel 文件 → 注册为可查询表（内部调用 register，返回 TableRef）。"""
    out = xlsx_to_table(content)
    return table_store.register(out["columns"], out["rows"], name=name, source=source)


def register_from_doc(doc_id: int) -> list[Any]:
    """把已解析文档中的 table 块批量注册为可查询表，返回 TableRef 列表。"""
    doc = db.get_document(doc_id)
    if doc is None:
        raise TableError("文档不存在", 404)
    if doc["status"] != "parsed":
        raise TableError("文档尚未完成解析，请稍后重试", 409)
    refs: list[Any] = []
    n = 0
    for chunk in db.list_chunks(doc_id):
        if chunk["kind"] != "table":
            continue
        n += 1
        try:
            out = markdown_to_table(chunk["content"] or "")
        except TableError:
            continue
        ref = table_store.register(
            out["columns"],
            out["rows"],
            name=f"{doc['name']} 表格{n}",
            source="doc",
            meta={"doc_id": doc_id, "chunk_id": chunk["id"], "page": chunk["page"]},
        )
        refs.append(ref)
    if not refs:
        raise TableError("文档中没有可查询的表格块", 400)
    logger.info("从文档 %s 注册表格 %s 张", doc_id, len(refs))
    return refs
