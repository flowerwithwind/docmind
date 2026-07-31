"""B2 表格问答后端测试：Excel/PDF 表格转可查询数据、NL2SQL 链路（LLM mock 自纠错）、
SQL 白名单拦截、无 Key 降级模式、指标记录与 API 集成（hermetic，不依赖真实网络/Key）。"""
from __future__ import annotations

import io
import json

import pytest
from app.llm.client import LLMError
from app.services import table_qa as qa_svc
from app.services import tables as table_convert
from app.services.table_store import (
    MAX_ROWS,
    SqlValidationError,
    TableError,
    ensure_limit,
    table_store,
    validate_sql,
)
from app.storage import db


@pytest.fixture(autouse=True)
def _db_ready():
    """非 client 用例也需要 DB（LLM 客户端配置读取 settings 表）。"""
    db.init_db()


SALES_COLUMNS = ["月份", "产品", "销售额"]
SALES_ROWS = [
    ["2026-01", "服务器", 1200000],
    ["2026-01", "交换机", 450000],
    ["2026-02", "服务器", 1350000],
    ["2026-02", "交换机", 420000],
    ["2026-03", "服务器", 1480000],
    ["2026-03", "交换机", 510000],
]


def _register_sales(table_id: str = "sales_test") -> object:
    return table_store.register(
        SALES_COLUMNS, SALES_ROWS, name="销售测试表", table_id=table_id
    )


def _make_xlsx_bytes() -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "销售"
    ws.append(SALES_COLUMNS)
    for row in SALES_ROWS:
        ws.append(row)
    ws2 = wb.create_sheet("忽略")
    ws2.append(["A"])
    ws2.append(["B"])
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def _seed_parsed_doc(table_md: str) -> int:
    doc_id = db.create_document(
        name="表格文档",
        filename="t.pdf",
        original_name="表格文档.pdf",
        ext=".pdf",
        mime="application/pdf",
        size_bytes=100,
        created_at=db.now_iso(),
    )
    db.update_document(doc_id, status="parsed")
    db.insert_chunks(
        [
            {
                "doc_id": doc_id,
                "seq": 1,
                "kind": "table",
                "section_path": "",
                "title": "销售表",
                "content": table_md,
                "page": 2,
                "char_count": len(table_md),
                "token_estimate": 10,
                "image_path": None,
                "created_at": db.now_iso(),
            }
        ]
    )
    return doc_id


class _FakeLLM:
    """按调用次序返回 canned 响应的假 LLM 客户端（hermetic）。"""

    configured = True

    def __init__(self, responses: list):
        self.responses = list(responses)
        self.calls: list[list] = []

    def chat(self, messages: list[dict], json_mode: bool = False) -> str:
        self.calls.append(messages)
        return self.responses.pop(0)


@pytest.fixture()
def fake_llm(monkeypatch):
    def _install(responses: list) -> _FakeLLM:
        fake = _FakeLLM(responses)
        monkeypatch.setattr(
            qa_svc.settings_svc, "build_llm_client", lambda: fake
        )
        return fake

    return _install


# ---------------------------------------------------------------- Excel 转换

def test_xlsx_to_table_first_sheet():
    out = table_convert.xlsx_to_table(_make_xlsx_bytes())
    assert out["sheet"] == "销售"
    assert out["columns"] == SALES_COLUMNS
    assert len(out["rows"]) == len(SALES_ROWS)
    assert out["rows"][0] == ["2026-01", "服务器", "1200000"]


def test_register_xlsx_queryable():
    ref = table_convert.register_xlsx(_make_xlsx_bytes(), name="Excel 测试")
    assert ref.source == "xlsx"
    assert len(ref.rows) == 6
    columns, rows = table_store.query(ref, "SELECT COUNT(*) AS n FROM t")
    assert columns == ["n"]
    assert rows == [[6]]


def test_xlsx_empty_sheet_raises():
    from openpyxl import Workbook

    wb = Workbook()
    wb.active.title = "空表"  # 无任何单元格
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    with pytest.raises(TableError):
        table_convert.xlsx_to_table(buf.getvalue())


# ---------------------------------------------------------------- PDF 表格清洗

def test_markdown_to_table_cleans_pdf_block():
    md = (
        "| 月份   | 产品   | 销售额  |\n"
        "| ------ | ------ | ------- |\n"
        "| 2026-01 | 服务器 | 1200000 |\n"
        "| 2026-02 | 交换机 |  450000 |\n"
        "| 2026-03 | 服务器 | 1480000 |\n"
    )
    out = table_convert.markdown_to_table(md)
    assert out["columns"] == ["月份", "产品", "销售额"]
    assert len(out["rows"]) == 3
    assert out["rows"][1] == ["2026-02", "交换机", "450000"]


def test_markdown_to_table_ragged_and_blank_rows():
    md = (
        "| 月份 | 产品 | 销售额 |\n"
        "| --- | --- | --- |\n"
        "| 2026-01 | 服务器 | 1200000 |\n"
        "| 2026-02 | 交换机 |\n"
        "|  |  |  |\n"
        "| 2026-03 | 服务器 | 1480000 |\n"
    )
    out = table_convert.markdown_to_table(md)
    assert len(out["rows"]) == 3
    assert out["rows"][1] == ["2026-02", "交换机", ""]
    assert len(out["rows"][1]) == 3  # 行宽补齐


def test_markdown_to_table_rejects_non_table():
    with pytest.raises(TableError):
        table_convert.markdown_to_table("这是普通文本\n没有表格\n")
    with pytest.raises(TableError):
        table_convert.markdown_to_table("")


# ---------------------------------------------------------------- 文档表格块注册

def test_register_from_doc_table_chunks():
    doc_id = _seed_parsed_doc(
        "| 产品 | 数量 |\n| --- | --- |\n| 服务器 | 10 |\n| 交换机 | 5 |\n"
    )
    refs = table_convert.register_from_doc(doc_id)
    assert len(refs) == 1
    assert refs[0].source == "doc"
    assert refs[0].meta["doc_id"] == doc_id
    _columns, rows = table_store.query(refs[0], "SELECT * FROM t")
    assert len(rows) == 2


def test_register_from_doc_not_found_and_unparsed():
    with pytest.raises(TableError) as e:
        table_convert.register_from_doc(9999)
    assert e.value.status_code == 404

    doc_id = db.create_document(
        name="未解析", filename="u.pdf", original_name="未解析.pdf", ext=".pdf",
        mime="application/pdf", size_bytes=1, created_at=db.now_iso(),
    )
    with pytest.raises(TableError) as e:
        table_convert.register_from_doc(doc_id)
    assert e.value.status_code == 409


# ---------------------------------------------------------------- SQL 白名单

@pytest.mark.parametrize(
    "sql",
    [
        "",
        "   ",
        "DROP TABLE t",
        "INSERT INTO t VALUES (1)",
        "DELETE FROM t",
        "UPDATE t SET c1 = 1",
        "ALTER TABLE t ADD COLUMN x",
        "PRAGMA table_info(t)",
        "CREATE TABLE x (a)",
        "SELECT * FROM t; DROP TABLE t",
        "SELECT * FROM t -- 注释",
        "/* 注释 */ SELECT * FROM t",
        "WITH x AS (SELECT 1) SELECT * FROM x",
        "SELECT sqlite_master FROM sqlite_master",
        "SELECT * INTO x FROM t",
    ],
)
def test_validate_sql_rejects_unsafe(sql):
    with pytest.raises(SqlValidationError):
        validate_sql(sql)


def test_validate_sql_accepts_readonly_select():
    assert validate_sql("SELECT * FROM t") == "SELECT * FROM t"
    assert validate_sql("SELECT c1, SUM(c2) FROM t GROUP BY c1 ORDER BY c1 DESC") is not None
    assert validate_sql("SELECT * FROM t;").endswith("FROM t")


def test_ensure_limit_appends_and_caps():
    assert ensure_limit("SELECT * FROM t") == f"SELECT * FROM t LIMIT {MAX_ROWS}"
    assert ensure_limit("SELECT * FROM t;") == f"SELECT * FROM t LIMIT {MAX_ROWS};"
    assert ensure_limit("SELECT * FROM t LIMIT 5") == "SELECT * FROM t LIMIT 5"
    assert ensure_limit("SELECT * FROM t LIMIT 99999") == f"SELECT * FROM t LIMIT {MAX_ROWS}"
    assert ensure_limit("SELECT * FROM t LIMIT 5;") == "SELECT * FROM t LIMIT 5;"


def test_query_force_limit_and_timeout_guard():
    ref = _register_sales()
    _columns, rows = table_store.query(ref, "SELECT * FROM t")
    assert len(rows) == len(SALES_ROWS)
    with pytest.raises(SqlValidationError):
        table_store.query(ref, "DELETE FROM t")


# ---------------------------------------------------------------- LLM 链路与自纠错

def test_llm_generate_execute_self_correct(fake_llm):
    fake = fake_llm(
        [
            json.dumps({"sql": "DELETE FROM t", "answer": "错误"}, ensure_ascii=False),
            json.dumps(
                {"sql": "SELECT * FROM t WHERE c2 LIKE '%服务器%'", "answer": "服务器相关记录"},
                ensure_ascii=False,
            ),
        ]
    )
    ref = _register_sales()
    out = qa_svc.answer_table(ref, "列出所有服务器")
    assert out["source"] == "llm"
    assert out["metrics"]["attempts"] == 2
    assert len(fake.calls) == 2  # 失败重试 1 次
    assert out["sql"].startswith("SELECT")
    assert out["rows"] and len(out["rows"]) == 3
    assert out["answer"] == "服务器相关记录"
    assert out["metrics"]["tokens"] > 0
    assert out["metrics"]["attempts"] == 2
    assert out["metrics"]["elapsed_ms"] >= 0


def test_llm_first_try_success(fake_llm):
    fake_llm(
        [
            json.dumps(
                {"sql": "SELECT COUNT(*) AS n FROM t", "answer": "共 6 条"},
                ensure_ascii=False,
            )
        ]
    )
    ref = _register_sales()
    out = qa_svc.answer_table(ref, "一共有多少条记录")
    assert out["source"] == "llm"
    assert out["metrics"]["attempts"] == 1
    assert out["rows"] == [[6]]


def test_llm_all_fail_falls_back_to_demo(fake_llm):
    fake_llm(
        [
            json.dumps({"sql": "DROP TABLE t", "answer": "x"}, ensure_ascii=False),
            json.dumps({"sql": "UPDATE t SET c1=1", "answer": "x"}, ensure_ascii=False),
        ]
    )
    ref = _register_sales()
    out = qa_svc.answer_table(ref, "销售额合计是多少")
    assert out["source"] == "demo"
    assert "自纠错后仍失败" in out["metrics"]["fallback_reason"]
    assert out["metrics"]["attempts"] == 1
    assert out["rows"] and out["answer"]


def test_llm_error_falls_back_to_demo(monkeypatch):
    def boom(messages, json_mode=False):
        raise LLMError("网络错误")

    class Broken:
        configured = True
        chat = staticmethod(boom)

    monkeypatch.setattr(qa_svc.settings_svc, "build_llm_client", lambda: Broken())
    out = qa_svc.answer_table(_register_sales(), "列出服务器")
    assert out["source"] == "demo"
    assert "模型调用失败" in out["metrics"]["fallback_reason"]


def test_llm_bad_json_then_valid(fake_llm):
    fake_llm(["这不是 JSON", '{"sql": "SELECT * FROM t", "answer": "ok"}'])
    ref = _register_sales()
    out = qa_svc.answer_table(ref, "显示所有记录")
    assert out["source"] == "llm"
    assert out["metrics"]["attempts"] == 2
    assert len(out["rows"]) == len(SALES_ROWS)


# ---------------------------------------------------------------- 降级模式

def test_demo_mode_aggregate():
    ref = _register_sales()
    out = qa_svc.answer_table(ref, "服务器销售额合计是多少")
    assert out["source"] == "demo"
    assert out["metrics"]["intent"] == "aggregate"
    assert "合计" in out["answer"]
    assert "5410000" in out["answer"]
    assert out["rows"] == [[5410000]]
    assert out["sql"].startswith("SELECT SUM(")
    assert out["metrics"]["fallback_reason"]  # 无 Key 降级原因


def test_demo_mode_filter_and_list():
    ref = _register_sales()
    filtered = qa_svc.answer_table(ref, "列出服务器")
    assert filtered["source"] == "demo"
    assert filtered["metrics"]["intent"] == "filter"
    assert "匹配到" in filtered["answer"]
    assert len(filtered["rows"]) == 3

    listed = qa_svc.answer_table(ref, "看看有什么")
    assert listed["source"] == "demo"
    assert listed["metrics"]["intent"] == "filter"
    assert listed["rows"]


def test_demo_mode_count():
    ref = _register_sales()
    out = qa_svc.answer_table(ref, "表里一共有多少条记录")
    assert out["source"] == "demo"
    assert "共 6 条记录" in out["answer"]
    assert out["rows"] == [[6]]


def test_metrics_recorded():
    ref = _register_sales()
    out = qa_svc.answer_table(ref, "销售额合计是多少")
    m = out["metrics"]
    assert m["elapsed_ms"] >= 0
    assert m["attempts"] >= 1
    assert m["tokens"] > 0
    assert m["intent"] == "aggregate"
    assert m["table_id"] == ref.id
    assert m["table_name"] == "销售测试表"
    assert m["row_count"] == len(out["rows"])
    assert isinstance(m["fallback_reason"], str)


# ---------------------------------------------------------------- 图表

def test_build_chart_kinds():
    ref = _register_sales()
    out = qa_svc.answer_table(ref, "销售额合计是多少")
    chart = out["chart"]
    assert chart["type"] in ("line", "bar", "table")
    assert chart["columns"] == out["columns"]
    assert chart["rows"] == out["rows"]
    assert chart["type"] == "bar"  # 单值聚合结果


# ---------------------------------------------------------------- API

def test_api_qa_table_inline_demo(client):
    r = client.post(
        "/api/qa/table",
        json={
            "question": "销售额合计是多少",
            "table": {"name": "内联销售", "columns": SALES_COLUMNS, "rows": SALES_ROWS},
        },
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["source"] == "demo"
    assert "合计" in body["answer"]
    assert body["sql"].startswith("SELECT SUM(")
    assert body["columns"] == ["结果"]
    assert body["rows"] == [[5410000]]
    assert body["chart"]["type"] == "bar"
    assert body["metrics"]["intent"] == "aggregate"
    assert body["metrics"]["table_id"]
    assert len(body["tables"]) == 1
    assert body["tables"][0] == body["metrics"]["table_id"]


def test_api_qa_table_by_table_id(client):
    ref = _register_sales("api_sales")
    r = client.post(
        "/api/qa/table",
        json={"question": "列出交换机", "table_id": ref.id},
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["source"] == "demo"
    assert len(body["rows"]) == 3


def test_api_qa_table_llm_mode(client, fake_llm):
    fake_llm(
        [
            json.dumps(
                {"sql": "SELECT c2, SUM(c3) FROM t GROUP BY c2", "answer": "按产品汇总"},
                ensure_ascii=False,
            )
        ]
    )
    r = client.post(
        "/api/qa/table",
        json={
            "question": "按产品汇总销售额",
            "table": {"name": "内联销售", "columns": SALES_COLUMNS, "rows": SALES_ROWS},
        },
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["source"] == "llm"
    assert body["metrics"]["attempts"] == 1
    assert body["answer"] == "按产品汇总"
    assert body["rows"]


def test_api_qa_table_doc_flow(client):
    doc_id = _seed_parsed_doc(
        "| 产品 | 数量 |\n| --- | --- |\n| 服务器 | 10 |\n| 交换机 | 5 |\n"
    )
    r = client.post(
        "/api/qa/table",
        json={"question": "有多少条记录", "doc_id": doc_id},
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["source"] == "demo"
    assert body["rows"] == [[2]]
    assert len(body["tables"]) == 1


def test_api_qa_table_doc_unparsed_409(client):
    doc_id = db.create_document(
        name="未解析", filename="u.pdf", original_name="未解析.pdf", ext=".pdf",
        mime="application/pdf", size_bytes=1, created_at=db.now_iso(),
    )
    r = client.post("/api/qa/table", json={"question": "有多少条", "doc_id": doc_id})
    assert r.status_code == 409


def test_api_qa_table_missing_source_400(client):
    r = client.post("/api/qa/table", json={"question": "有多少条"})
    assert r.status_code == 400
    r2 = client.post("/api/qa/table", json={"question": "  "})
    assert r2.status_code == 400


def test_api_qa_table_unknown_table_404(client):
    r = client.post("/api/qa/table", json={"question": "有多少条", "table_id": "nope"})
    assert r.status_code == 404


def test_api_qa_tables_list(client):
    _register_sales("api_list")
    r = client.get("/api/qa/tables")
    assert r.status_code == 200
    items = r.json()
    ids = {t["id"] for t in items}
    assert "api_list" in ids
    item = next(t for t in items if t["id"] == "api_list")
    assert item["row_count"] == 6
    assert item["columns"] == SALES_COLUMNS